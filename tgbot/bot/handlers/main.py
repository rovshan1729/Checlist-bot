from aiogram import types
from aiogram.contrib.middlewares.logging import LoggingMiddleware
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Regexp

from tgbot.bot.states.states import (
    Intro, 
    FirstSection, 
    SecondSection, 
    ThirdSection,
    FourSection,
    FifthSection,
    SixSection,
    SevenSection,
    TextQuestionSection,
    )
from tgbot.bot.loader import dp
from tgbot.models import Section, SectionNumberChoices 
from tgbot.bot.keyboards.inline import get_back_keyboard

import re
import math

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

dp.middleware.setup(LoggingMiddleware())


@dp.message_handler(commands=['restart'], state="*")
async def restart_command_handler(message: types.Message, state: FSMContext):
    await state.finish()
    await message.answer("Теперь вы можете начать с начало.") 
    await start_handler(message) 


@dp.callback_query_handler(lambda c: c.data == 'back', state='*')
async def process_back_button(callback_query: types.CallbackQuery, state: FSMContext):
    await callback_query.answer()
    
    current_state = await state.get_state()
    data = await state.get_data()
    total_ball = data.get('total_ball', 0)
    questions = data.get('questions', [])
    
    if current_state == Intro.name.state:
        previous_state = Intro.branch
        text = f"<strong>Здравствуйте! Какой филиал вы проверяете?</strong>"
    
    elif current_state == FirstSection.s1q1.state:
        previous_state = Intro.name
        text = f"<strong>Введите полное имя проверяющего.</strong>"

    elif current_state == FirstSection.s1q2.state:
        previous_state = FirstSection.s1q1
        total_ball -= int(data.get('s1q1', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[0].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FirstSection.s1q3.state:
        previous_state = FirstSection.s1q2
        total_ball -= int(data.get('s1q2', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[1].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FirstSection.s1q4.state:
        previous_state = FirstSection.s1q3
        total_ball -= int(data.get('s1q3', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[2].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FirstSection.s1q5.state:
        previous_state = FirstSection.s1q4
        total_ball -= int(data.get('s1q4', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[3].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FirstSection.s1q6.state:
        previous_state = FirstSection.s1q5
        total_ball -= int(data.get('s1q5', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[4].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FirstSection.s1q7.state:
        previous_state = FirstSection.s1q6
        total_ball -= int(data.get('s1q6', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[5].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FirstSection.s1q8.state:
        previous_state = FirstSection.s1q7
        total_ball -= int(data.get('s1q7', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[6].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FirstSection.s1q9.state:
        previous_state = FirstSection.s1q8
        total_ball -= int(data.get('s1q8', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[7].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FirstSection.s1q10.state:
        previous_state = FirstSection.s1q9
        total_ball -= int(data.get('s1q9', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[8].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FirstSection.s1q11.state:
        previous_state = FirstSection.s1q10
        total_ball -= int(data.get('s1q10', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[9].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FirstSection.s1q12.state:
        previous_state = FirstSection.s1q11
        total_ball -= int(data.get('s1q11', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[10].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FirstSection.s1q13.state:
        previous_state = FirstSection.s1q12
        total_ball -= int(data.get('s1q12', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[11].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FirstSection.s1q14.state:
        previous_state = FirstSection.s1q13
        total_ball -= int(data.get('s1q13', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[12].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
        """
            Next section 
        """
    elif current_state == SecondSection.s2q1.state:
        previous_state = FirstSection.s1q14
        total_ball = int(data.get('s1q13', 0))
        section = Section.objects.filter(type=SectionNumberChoices.FIRST).first()
        questions = list(section.section_questions.order_by('order'))
        await state.update_data(questions=questions)
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[13].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SecondSection.s2q2.state:
        previous_state = SecondSection.s2q1
        total_ball -= int(data.get('s2q1', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[0].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SecondSection.s2q3.state:
        previous_state = SecondSection.s2q2
        total_ball -= int(data.get('s2q2', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[1].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SecondSection.s2q4.state:
        previous_state = SecondSection.s2q3
        total_ball -= int(data.get('s2q3', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[2].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SecondSection.s2q5.state:
        previous_state = SecondSection.s2q4
        total_ball -= int(data.get('s2q4', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[3].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SecondSection.s2q6.state:
        previous_state = SecondSection.s2q5
        total_ball -= int(data.get('s2q5', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[4].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SecondSection.s2q7.state:
        previous_state = SecondSection.s2q6
        total_ball -= int(data.get('s2q6', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[5].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SecondSection.s2q8.state:
        previous_state = SecondSection.s2q7
        total_ball -= int(data.get('s2q7', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[6].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SecondSection.s2q9.state:
        previous_state = SecondSection.s2q8
        total_ball -= int(data.get('s2q8', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[7].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SecondSection.s2q10.state:
        previous_state = SecondSection.s2q9
        total_ball -= int(data.get('s2q9', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[8].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SecondSection.s2q11.state:
        previous_state = SecondSection.s2q10
        total_ball -= int(data.get('s2q10', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[9].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SecondSection.s2q12.state:
        previous_state = SecondSection.s2q11
        total_ball -= int(data.get('s2q11', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[10].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SecondSection.s2q13.state:
        previous_state = SecondSection.s2q12
        total_ball -= int(data.get('s2q12', 0))
        text = (
        f'<strong>{questions[0].section.title}</strong>\n\n'
        f'{questions[11].title}\n\n'
        'Оцените, пожалуйста, от 0 до 100 баллов.'
        )
         
        """
        Next Section
        """
        
    elif current_state == ThirdSection.s3q1.state:
        previous_state = SecondSection.s2q13
        total_ball = int(data.get('s2q12', 0))
        section = Section.objects.filter(type=SectionNumberChoices.SECOND).first()
        questions = list(section.section_questions.order_by('order'))
        await state.update_data(questions=questions)
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[12].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == ThirdSection.s3q2.state:
        previous_state = ThirdSection.s3q1
        total_ball -= int(data.get('s3q1', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[0].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == ThirdSection.s3q3.state:
        previous_state = ThirdSection.s3q2
        total_ball -= int(data.get('s3q2', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[1].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == ThirdSection.s3q4.state:
        previous_state = ThirdSection.s3q3
        total_ball -= int(data.get('s3q3', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[2].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == ThirdSection.s3q5.state:
        previous_state = ThirdSection.s3q4
        total_ball -= int(data.get('s3q4', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[3].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == ThirdSection.s3q6.state:
        previous_state = ThirdSection.s3q5
        total_ball -= int(data.get('s3q5', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[4].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == ThirdSection.s3q7.state:
        previous_state = ThirdSection.s3q6
        total_ball -= int(data.get('s3q6', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[5].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == ThirdSection.s3q8.state:
        previous_state = ThirdSection.s3q7
        total_ball -= int(data.get('s3q7', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[6].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == ThirdSection.s3q9.state:
        previous_state = ThirdSection.s3q8
        total_ball -= int(data.get('s3q8', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[7].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == ThirdSection.s3q10.state:
        previous_state = ThirdSection.s3q9
        total_ball -= int(data.get('s3q9', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[8].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == ThirdSection.s3q11.state:
        previous_state = ThirdSection.s3q10
        total_ball -= int(data.get('s3q10', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[9].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == ThirdSection.s3q12.state:
        previous_state = ThirdSection.s3q11
        total_ball -= int(data.get('s3q11', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[10].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == ThirdSection.s3q13.state:
        previous_state = ThirdSection.s3q12
        total_ball -= int(data.get('s3q12', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[11].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == ThirdSection.s3q14.state:
        previous_state = ThirdSection.s3q13
        total_ball -= int(data.get('s3q13', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[12].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == ThirdSection.s3q15.state:
        previous_state = ThirdSection.s3q14
        total_ball -= int(data.get('s3q14', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[13].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == ThirdSection.s3q16.state:
        previous_state = ThirdSection.s3q15
        total_ball -= int(data.get('s3q15', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[14].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == ThirdSection.s3q17.state:
        previous_state = ThirdSection.s3q16
        total_ball -= int(data.get('s3q16', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[15].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == ThirdSection.s3q18.state:
        previous_state = ThirdSection.s3q17
        total_ball -= int(data.get('s3q17', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[16].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == ThirdSection.s3q19.state:
        previous_state = ThirdSection.s3q18
        total_ball -= int(data.get('s3q18', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[17].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == ThirdSection.s3q20.state:
        previous_state = ThirdSection.s3q19
        total_ball -= int(data.get('s3q19', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[18].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == ThirdSection.s3q21.state:
        previous_state = ThirdSection.s3q20
        total_ball -= int(data.get('s3q20', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[19].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
        """
        Next section
        """
    elif current_state == FourSection.s4q1.state:
        previous_state = ThirdSection.s3q21
        total_ball = int(data.get('s3q20', 0))
        section = Section.objects.filter(type=SectionNumberChoices.THIRD).first()
        questions = list(section.section_questions.order_by('order'))
        await state.update_data(questions=questions)
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[20].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FourSection.s4q2.state:
        previous_state = FourSection.s4q1
        total_ball -= int(data.get('s4q1', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[0].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FourSection.s4q3.state:
        previous_state = FourSection.s4q2
        total_ball -= int(data.get('s4q2', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[1].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FourSection.s4q4.state:
        previous_state = FourSection.s4q3
        total_ball -= int(data.get('s4q3', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[2].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FourSection.s4q5.state:
        previous_state = FourSection.s4q4
        total_ball -= int(data.get('s4q4', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[3].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FourSection.s4q6.state:
        previous_state = FourSection.s4q5
        total_ball -= int(data.get('s4q5', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[4].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FourSection.s4q7.state:
        previous_state = FourSection.s4q6
        total_ball -= int(data.get('s4q6', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[5].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FourSection.s4q8.state:
        previous_state = FourSection.s4q7
        total_ball -= int(data.get('s4q7', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[6].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FourSection.s4q9.state:
        previous_state = FourSection.s4q8
        total_ball -= int(data.get('s4q8', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[7].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FourSection.s4q10.state:
        previous_state = FourSection.s4q9
        total_ball -= int(data.get('s4q9', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[8].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FourSection.s4q11.state:
        previous_state = FourSection.s4q10
        total_ball -= int(data.get('s4q10', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[9].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FourSection.s4q12.state:
        previous_state = FourSection.s4q11
        total_ball -= int(data.get('s4q11', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[10].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FourSection.s4q13.state:
        previous_state = FourSection.s4q12
        total_ball -= int(data.get('s4q12', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[11].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FourSection.s4q14.state:
        previous_state = FourSection.s4q13
        total_ball -= int(data.get('s4q13', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[12].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
        """
        Next Section
        """
        
    elif current_state == FifthSection.s5q1.state:
        previous_state = FourSection.s4q14
        total_ball = int(data.get('s4q13', 0))
        section = Section.objects.filter(type=SectionNumberChoices.FOUR).first()
        questions = list(section.section_questions.order_by('order'))
        await state.update_data("questions")
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[13].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
    
    elif current_state == FifthSection.s5q2.state:
        previous_state = FifthSection.s5q1
        total_ball -= int(data.get('s5q1', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[0].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FifthSection.s5q3.state:
        previous_state = FifthSection.s5q2
        total_ball -= int(data.get('s5q2', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[1].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FifthSection.s5q4.state:
        previous_state = FifthSection.s5q3
        total_ball -= int(data.get('s5q3', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[2].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FifthSection.s5q5.state:
        previous_state = FifthSection.s5q4
        total_ball -= int(data.get('s5q4', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[3].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FifthSection.s5q6.state:
        previous_state = FifthSection.s5q5
        total_ball -= int(data.get('s5q5', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[4].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FifthSection.s5q7.state:
        previous_state = FifthSection.s5q6
        total_ball -= int(data.get('s5q6', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[5].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FifthSection.s5q8.state:
        previous_state = FifthSection.s5q7
        total_ball -= int(data.get('s5q7', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[6].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FifthSection.s5q9.state:
        previous_state = FifthSection.s5q8
        total_ball -= int(data.get('s5q8', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[7].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FifthSection.s5q10.state:
        previous_state = FifthSection.s5q9
        total_ball -= int(data.get('s5q9', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[8].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FifthSection.s5q11.state:
        previous_state = FifthSection.s5q10
        total_ball -= int(data.get('s5q10', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[9].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FifthSection.s5q12.state:
        previous_state = FifthSection.s5q11
        total_ball -= int(data.get('s5q11', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[10].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FifthSection.s5q13.state:
        previous_state = FifthSection.s5q12
        total_ball -= int(data.get('s5q12', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[11].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FifthSection.s5q14.state:
        previous_state = FifthSection.s5q13
        total_ball -= int(data.get('s5q13', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[12].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FifthSection.s5q15.state:
        previous_state = FifthSection.s5q14
        total_ball -= int(data.get('s5q14', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[13].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FifthSection.s5q16.state:
        previous_state = FifthSection.s5q15
        total_ball -= int(data.get('s5q15', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[14].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FifthSection.s5q17.state:
        previous_state = FifthSection.s5q16
        total_ball -= int(data.get('s5q16', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[15].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FifthSection.s5q18.state:
        previous_state = FifthSection.s5q17
        total_ball -= int(data.get('s5q17', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[16].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FifthSection.s5q19.state:
        previous_state = FifthSection.s5q18
        total_ball -= int(data.get('s5q18', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[17].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == FifthSection.s5q20.state:
        previous_state = FifthSection.s5q19
        total_ball -= int(data.get('s5q19', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[18].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
        """
        Next Section
        """
        
    elif current_state == SixSection.s6q1.state:
        previous_state = FifthSection.s5q20
        total_ball = int(data.get('s5q19', 0))
        section = Section.objects.filter(type=SectionNumberChoices.FIVE).first()
        questions = list(section.section_questions.order_by('order'))
        await state.update_data(questions=questions)
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[19].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SixSection.s6q2.state:
        previous_state = SixSection.s6q1
        total_ball -= int(data.get('s6q1', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[0].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SixSection.s6q3.state:
        previous_state = SixSection.s6q2
        total_ball -= int(data.get('s6q2', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[1].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SixSection.s6q4.state:
        previous_state = SixSection.s6q3
        total_ball -= int(data.get('s6q3', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[2].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SixSection.s6q5.state:
        previous_state = SixSection.s6q4
        total_ball -= int(data.get('s6q4', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[3].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SixSection.s6q6.state:
        previous_state = SixSection.s6q5
        total_ball -= int(data.get('s6q5', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[4].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SixSection.s6q7.state:
        previous_state = SixSection.s6q6
        total_ball -= int(data.get('s6q6', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[5].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SixSection.s6q8.state:
        previous_state = SixSection.s6q7
        total_ball -= int(data.get('s6q7', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[6].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SixSection.s6q9.state:
        previous_state = SixSection.s6q8
        total_ball -= int(data.get('s6q8', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[7].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
        """
        Next Section
        """
        
    elif current_state == SevenSection.s7q1.state:
        previous_state = SixSection.s6q9
        total_ball = int(data.get('s6q8', 0))
        section = Section.objects.filter(type=SectionNumberChoices.SIX).first()
        questions = list(section.section_questions.order_by('order'))
        await state.update_data(questions=questions)
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[8].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SevenSection.s7q2.state:
        previous_state = SevenSection.s7q1
        total_ball -= int(data.get('s7q1', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[0].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SevenSection.s7q3.state:
        previous_state = SevenSection.s7q2
        total_ball -= int(data.get('s7q2', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[1].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SevenSection.s7q4.state:
        previous_state = SevenSection.s7q3
        total_ball -= int(data.get('s7q3', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[2].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SevenSection.s7q5.state:
        previous_state = SevenSection.s7q4
        total_ball -= int(data.get('s7q4', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[3].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SevenSection.s7q6.state:
        previous_state = SevenSection.s7q5
        total_ball -= int(data.get('s7q5', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[4].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SevenSection.s7q7.state:
        previous_state = SevenSection.s7q6
        total_ball -= int(data.get('s7q6', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[5].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SevenSection.s7q8.state:
        previous_state = SevenSection.s7q7
        total_ball -= int(data.get('s7q7', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[6].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SevenSection.s7q9.state:
        previous_state = SevenSection.s7q8
        total_ball -= int(data.get('s7q8', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[7].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SevenSection.s7q10.state:
        previous_state = SevenSection.s7q9
        total_ball -= int(data.get('s7q9', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[8].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == SevenSection.s7q11.state:
        previous_state = SevenSection.s7q10
        total_ball -= int(data.get('s7q10', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[9].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        """
        Next Section
        """
    elif current_state == TextQuestionSection.t1q1.state:
        previous_state = SevenSection.s7q11
        section = Section.objects.filter(type=SectionNumberChoices.SEVEN).first()
        questions = list(section.section_questions.order_by('order'))
        await state.update_data(questions=questions)
        total_ball = int(data.get('s7q10', 0))
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[10].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == TextQuestionSection.t1q2.state:
        previous_state = TextQuestionSection.t1q1
        t1q1 = ""
        await state.update_data(t1q1=t1q1)
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[0].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == TextQuestionSection.t1q3.state:
        previous_state = TextQuestionSection.t1q2
        section = Section.objects.filter(type=SectionNumberChoices.TEXT_QUESTION).first()
        questions = list(section.section_questions.order_by('order'))
        await state.update_data(questions=questions)
        t1q2 = ""
        await state.update_data(t1q2=t1q2)
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[1].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    elif current_state == TextQuestionSection.t1q4.state:
        previous_state = TextQuestionSection.t1q3
        t1q3 = ""
        await state.update_data(t1q3=t1q3)
        text = f'<strong>{questions[0].section.title}</strong>\n\n{questions[0].title}\n\nОцените, пожалуйста, от 0 до 100 баллов.'
        
    else:
        await callback_query.message.answer("Вы вернулись в главное меню.\nНажмите /start")
        await state.finish()
        return

    if callback_query.message.text != text or callback_query.message.reply_markup != get_back_keyboard():
        await callback_query.message.edit_text(text, reply_markup=get_back_keyboard(),
                                               parse_mode='HTML')

    await state.update_data(total_ball=total_ball)
    await previous_state.set()


@dp.message_handler(commands=['start'])
async def start_handler(message: types.Message):
    await message.answer("<strong>Здравствуйте! Какой филиал вы проверяете?</strong>",
                         reply_markup=get_back_keyboard(),
                         parse_mode='HTML')
    await Intro.branch.set()

@dp.message_handler(Regexp(r'^[A-Za-zА-Яа-яЁё\s]+$'), state=Intro.branch)
async def branch_handler(message: types.Message, state: FSMContext):
    await state.update_data(branch=message.text)
    await message.answer("<strong>Введите полное имя проверяющего.</strong>",
                         reply_markup=get_back_keyboard(),
                         parse_mode='HTML')
    await Intro.name.set()

@dp.message_handler(lambda message: not re.match(r'^[A-Za-zА-Яа-яЁё\s]+$', message.text), state=Intro.branch)
async def invalid_branch_handler(message: types.Message):
    await message.answer("Ошибка! Пожалуйста, введите название филиала, используя только буквы.")

@dp.message_handler(lambda message: len(message.text.split()) in [2, 3], state=Intro.name)
async def name_handler(message: types.Message, state: FSMContext):
    await state.update_data(name=message.text)
    await message.answer("<strong>Введите время проверки (в формате ЧЧ:ММ).</strong>",
                         reply_markup=get_back_keyboard(),
                         parse_mode='HTML')
    await Intro.time.set()

@dp.message_handler(lambda message: len(message.text.split()) not in [2, 3], state=Intro.name)
async def invalid_name_handler(message: types.Message):
    await message.answer("Ошибка! Имя должно состоять из 2 или 3 слов.")
    

@dp.message_handler(lambda message: re.match(r'^\d{2}:\d{2}$', message.text), state=Intro.time)
async def time_handler(message: types.Message, state: FSMContext):
    await state.update_data(time=message.text)
    await message.answer("<strong>Введите дату проверки (в формате ДД.ММ.ГГГГ).</strong>",
                         reply_markup=get_back_keyboard(),
                         parse_mode='HTML')
    await Intro.date.set()

@dp.message_handler(lambda message: not re.match(r'^\d{2}:\d{2}$', message.text), state=Intro.time)
async def invalid_time_handler(message: types.Message):
    await message.answer("Ошибка! Пожалуйста, введите корректное время в формате ЧЧ:ММ.")

@dp.message_handler(lambda message: re.match(r'^\d{2}\.\d{2}\.\d{4}$', message.text), state=Intro.date)
async def date_handler(message: types.Message, state: FSMContext):
    await state.update_data(date=message.text)
    
    sections = Section.objects.filter(type=SectionNumberChoices.FIRST).first()
    questions = list(sections.section_questions.order_by('order'))
    await state.update_data(questions=questions)
    
    await message.answer(f'<strong>{sections.title}</strong>\n\n{questions[0].title}\n\n'
                         f'Оцените, пожалуйста, от 0 до 100 баллов.',
                         reply_markup=get_back_keyboard(),
                         parse_mode='HTML')
    await FirstSection.s1q1.set()

@dp.message_handler(lambda message: not re.match(r'^\d{2}\.\d{2}\.\d{4}$', message.text), state=Intro.date)
async def invalid_date_handler(message: types.Message):
    await message.answer("Ошибка! Пожалуйста, введите корректную дату в формате ДД.ММ.ГГГГ.")


@dp.message_handler(state=FirstSection.s1q1)
async def s1q1_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            total_ball = int(message.text)
            await state.update_data(total_ball=total_ball, s1q1=int(message.text))
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[1].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await FirstSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")

@dp.message_handler(state=FirstSection.s1q2)
async def s1q2_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s1q2=int(message.text))
            questions = data.get('questions', [])

            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[2].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await FirstSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")


@dp.message_handler(state=FirstSection.s1q3)
async def s1q3_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s1q3=int(message.text))
            questions = data.get('questions', [])

            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[3].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await FirstSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")


@dp.message_handler(state=FirstSection.s1q4)
async def s1q4_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s1q4=int(message.text))
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[4].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await FirstSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")

@dp.message_handler(state=FirstSection.s1q5)
async def s1q5_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s1q5=int(message.text))
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[5].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await FirstSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")


@dp.message_handler(state=FirstSection.s1q6)
async def s1q6_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s1q6=int(message.text))
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[6].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            await FirstSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")


@dp.message_handler(state=FirstSection.s1q7)
async def s1q7_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s1q7=int(message.text))
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[7].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await FirstSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")


@dp.message_handler(state=FirstSection.s1q8)
async def s1q8_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s1q8=int(message.text))
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[8].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await FirstSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")


@dp.message_handler(state=FirstSection.s1q9)
async def s1q9_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s1q9=int(message.text))
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[9].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await FirstSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")


@dp.message_handler(state=FirstSection.s1q10)
async def s1q10_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s1q10=int(message.text))
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[10].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await FirstSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")

@dp.message_handler(state=FirstSection.s1q11)
async def s1q11_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s1q11=int(message.text))
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[11].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await FirstSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")


@dp.message_handler(state=FirstSection.s1q12)
async def s1q12_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s1q12=int(message.text))
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[12].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await FirstSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")

@dp.message_handler(state=FirstSection.s1q13)
async def s1q13_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s1q13=int(message.text))
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[13].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await FirstSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")

@dp.message_handler(state=FirstSection.s1q14)
async def s1q14_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s1q14=int(message.text))
            
            section = Section.objects.filter(type=SectionNumberChoices.FIRST).first()
            total_questions = section.total_questions
            
            first_section_result = (total_ball // total_questions) * 0.2
            await state.update_data(first_section_result=first_section_result)
            
            await message.answer(f'<strong>Вы оценили первый {section.title}</strong>\n'
                     f'Результат {first_section_result}% из 20% общего\n\n'
                     f'Чтобы продолжить оценивание и перейти к следующему разделу, нажмите на /next')
            
            await FirstSection.next()
            
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")    
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
    

@dp.message_handler(state=FirstSection.final_s1)
async def final_s1_handler(message: types.Message, state: FSMContext):
    if message.text == "/next":

        total_ball = 0
        section = Section.objects.filter(type=SectionNumberChoices.SECOND).first()
        questions = list(section.section_questions.order_by('order'))
        await state.update_data(questions=questions, total_ball=total_ball)
        
        await message.answer(f'<strong>Вы перешли на другую секцию</strong>')
        await message.answer(
            f'<strong>{questions[0].section.title}</strong>\n\n{questions[0].title}\n\n'
            f'Оцените, пожалуйста, от 0 до 100 баллов.',
            reply_markup=get_back_keyboard(),
            parse_mode='HTML'
        )
        await SecondSection.s2q1.set()
    else:
        await message.answer("Нажмите /next чтобы продолжить")


@dp.message_handler(state=SecondSection.s2q1)
async def s2q1_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            await state.update_data(s2q1=int(message.text))
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball)
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[1].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await SecondSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SecondSection.s2q2)
async def s2q2_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            await state.update_data(s2q2=int(message.text))
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball)
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[2].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await SecondSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SecondSection.s2q3)
async def s2q3_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            await state.update_data(s2q3=int(message.text))
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball)
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[3].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await SecondSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SecondSection.s2q4)
async def s2q4_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            await state.update_data(s2q4=int(message.text))
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball)
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[4].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await SecondSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SecondSection.s2q5)
async def s2q5_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            await state.update_data(s2q5=int(message.text))
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball)
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[5].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await SecondSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SecondSection.s2q6)
async def s2q6_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            await state.update_data(s2q6=int(message.text))
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball)
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[6].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await SecondSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SecondSection.s2q7)
async def s2q7_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            await state.update_data(s2q7=int(message.text))
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball)
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[7].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await SecondSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")


@dp.message_handler(state=SecondSection.s2q8)
async def s2q8_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            await state.update_data(s2q8=int(message.text))
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball)
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[8].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await SecondSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SecondSection.s2q9)
async def s2q9_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s2q9=int(message.text))
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[9].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await SecondSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SecondSection.s2q10)
async def s2q10_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            await state.update_data(s2q10=int(message.text))
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball)
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[10].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await SecondSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SecondSection.s2q11)
async def s2q11_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            await state.update_data(s2q11=int(message.text))
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball)
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[11].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await SecondSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SecondSection.s2q12)
async def s2q12_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s2q12=int(message.text))
            questions = data.get('questions', [])
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[12].title}\n\n'
                                f'Оцените, пожалуйста, от 0 до 100 баллов.',
                                reply_markup=get_back_keyboard(),
                                parse_mode='HTML')
            
            await SecondSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SecondSection.s2q13)
async def s2q13_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s2q13=int(message.text))
            
            section = Section.objects.filter(type=SectionNumberChoices.SECOND).first()
            total_questions = section.total_questions
            
            second_section_result = (total_ball // total_questions) * 0.2
            await state.update_data(second_section_result=second_section_result)
           
            await message.answer(f'<strong>Вы оценили первый {section.title}</strong>\n'
                     f'Результат {second_section_result}% из 20% общего\n\n'
                     f'Чтобы продолжить оценивание и перейти к следующему разделу, нажмите на /next')
            await SecondSection.next()
            
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
    
@dp.message_handler(state=SecondSection.final_s2)
async def final_s2_handler(message: types.Message, state: FSMContext):
    if message.text == "/next":
        total_ball = 0
        
        section = Section.objects.filter(type=SectionNumberChoices.THIRD).first()
        questions = list(section.section_questions.order_by('order'))
        await state.update_data(questions=questions, total_ball=total_ball)
        
        await message.answer(f'<strong>Вы перешли на другую секцию</strong>')
        await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[0].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
        
        await ThirdSection.s3q1.set()
    else:
        await message.answer("Нажмите /next чтобы продолжить")    

    
@dp.message_handler(state=ThirdSection.s3q1)
async def s3q1_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s3q1=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[1].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await ThirdSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        

@dp.message_handler(state=ThirdSection.s3q2)
async def s3q2_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s3q2=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[2].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await ThirdSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=ThirdSection.s3q3)
async def s3q3_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s3q3=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[3].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await ThirdSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=ThirdSection.s3q4)
async def s3q4_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s3q4=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[4].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await ThirdSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=ThirdSection.s3q5)
async def s3q5_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s3q5=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[5].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await ThirdSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=ThirdSection.s3q6)
async def s3q6_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s3q6=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[6].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await ThirdSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=ThirdSection.s3q7)
async def s3q7_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s3q7=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[7].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await ThirdSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        

@dp.message_handler(state=ThirdSection.s3q8)
async def s3q8_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s3q8=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[8].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await ThirdSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=ThirdSection.s3q9)
async def s3q9_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s3q9=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[9].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await ThirdSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=ThirdSection.s3q10)
async def s3q10_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s3q10=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[10].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await ThirdSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=ThirdSection.s3q11)
async def s3q11_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s3q11=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[11].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await ThirdSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=ThirdSection.s3q12)
async def s3q12_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s3q12=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[12].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await ThirdSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=ThirdSection.s3q13)
async def s3q13_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s3q13=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[13].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await ThirdSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=ThirdSection.s3q14)
async def s3q14_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s3q14=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[14].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await ThirdSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=ThirdSection.s3q15)
async def s3q15_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s3q15=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[15].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await ThirdSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=ThirdSection.s3q16)
async def s3q16_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s3q16=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[16].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await ThirdSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=ThirdSection.s3q17)
async def s3q17_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s3q17=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[17].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await ThirdSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=ThirdSection.s3q18)
async def s3q18_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s3q18=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[18].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await ThirdSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=ThirdSection.s3q19)
async def s3q19_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s3q19=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[19].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await ThirdSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        

@dp.message_handler(state=ThirdSection.s3q20)
async def s3q20_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s3q20=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[20].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await ThirdSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=ThirdSection.s3q21)
async def s3q21_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s3q21=int(message.text))
            
            section = Section.objects.filter(type=SectionNumberChoices.THIRD).first()
            total_questions = section.total_questions
            
            third_section_result = (total_ball // total_questions) * 0.2
            await state.update_data(third_section_result=third_section_result)
            
            await message.answer(f'<strong>Вы оценили первый {section.title}</strong>\n'
                     f'Результат {third_section_result}% из 20% общего\n\n'
                     f'Чтобы продолжить оценивание и перейти к следующему разделу, нажмите на /next')
            
            await ThirdSection.next()
            
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")    
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=ThirdSection.final_s3)
async def final_s3_handler(message: types.Message, state: FSMContext):
    if message.text == "/next":

        total_ball = 0
        section = Section.objects.filter(type=SectionNumberChoices.FOUR).first()
        questions = list(section.section_questions.order_by('order'))
        await state.update_data(questions=questions, total_ball=total_ball)
        
        await message.answer(f'<strong>Вы перешли на другую секцию</strong>')
        await message.answer(
            f'<strong>{questions[0].section.title}</strong>\n\n{questions[0].title}\n\n'
            f'Оцените, пожалуйста, от 0 до 100 баллов.',
            reply_markup=get_back_keyboard(),
            parse_mode='HTML'
        )
        await FourSection.s4q1.set()
    else:
        await message.answer("Нажмите /next чтобы продолжить")
        

@dp.message_handler(state=FourSection.s4q1)
async def s4q1_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s4q1=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[1].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FourSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
        
@dp.message_handler(state=FourSection.s4q2)
async def s4q2_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s4q2=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[2].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FourSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FourSection.s4q3)
async def s4q3_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s4q3=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[3].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FourSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FourSection.s4q4)
async def s4q4_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s4q4=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[4].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FourSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FourSection.s4q5)
async def s4q5_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s4q5=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[5].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FourSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FourSection.s4q6)
async def s4q6_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s4q6=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[6].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FourSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FourSection.s4q7)
async def s4q7_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s4q7=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[7].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FourSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FourSection.s4q8)
async def s4q8_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s4q8=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[8].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FourSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FourSection.s4q9)
async def s4q9_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s4q9=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[9].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FourSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")

        
@dp.message_handler(state=FourSection.s4q10)
async def s4q10_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s4q10=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[10].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FourSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FourSection.s4q11)
async def s4q11_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s4q11=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[11].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FourSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FourSection.s4q12)
async def s4q12_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s4q12=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[12].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FourSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        

@dp.message_handler(state=FourSection.s4q13)
async def s4q13_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s4q13=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[13].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FourSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FourSection.s4q14)
async def s4q14_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s4q14=int(message.text))
            
            section = Section.objects.filter(type=SectionNumberChoices.FOUR).first()
            total_questions = section.total_questions
            
            four_section_result = (total_ball // total_questions) * 0.1
            await state.update_data(four_section_result=four_section_result)
            
            await message.answer(f'<strong>Вы оценили первый {section.title}</strong>\n'
                     f'Результат {four_section_result}% из 10% общего\n\n'
                     f'Чтобы продолжить оценивание и перейти к следующему разделу, нажмите на /next')
            
            await FourSection.next()
            
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")    
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FourSection.final_s4)
async def final_s4_handler(message: types.Message, state: FSMContext):
    if message.text == "/next":
        total_ball = 0
        
        section = Section.objects.filter(type=SectionNumberChoices.FIVE).first()
        questions = list(section.section_questions.order_by('order'))
        await state.update_data(questions=questions, total_ball=total_ball)
        
        await message.answer(f'<strong>Вы перешли на другую секцию</strong>')
        await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[0].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
        
        await FifthSection.s5q1.set()
    else:
        await message.answer("Нажмите /next чтобы продолжить") 
        
@dp.message_handler(state=FifthSection.s5q1)
async def s5q1_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s5q1=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[1].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FifthSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")   
        
        
@dp.message_handler(state=FifthSection.s5q1)
async def s5q1_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s5q1=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[1].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FifthSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FifthSection.s5q2)
async def s5q2_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s5q2=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[2].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FifthSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FifthSection.s5q3)
async def s5q3_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s5q3=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[3].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FifthSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FifthSection.s5q4)
async def s5q4_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s5q4=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[4].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FifthSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FifthSection.s5q5)
async def s5q5_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s5q5=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[5].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FifthSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FifthSection.s5q6)
async def s5q6_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s5q6=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[6].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FifthSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FifthSection.s5q7)
async def s5q7_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s5q7=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[7].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FifthSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FifthSection.s5q8)
async def s5q8_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s5q8=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[8].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FifthSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FifthSection.s5q9)
async def s5q9_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s5q9=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[9].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FifthSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FifthSection.s5q10)
async def s5q10_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s5q10=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[10].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FifthSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FifthSection.s5q11)
async def s5q11_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s5q11=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[11].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FifthSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FifthSection.s5q12)
async def s5q12_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s5q12=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[12].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FifthSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FifthSection.s5q13)
async def s5q13_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s5q13=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[13].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FifthSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FifthSection.s5q14)
async def s5q14_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s5q14=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[14].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FifthSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FifthSection.s5q15)
async def s5q15_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s5q15=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[15].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FifthSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FifthSection.s5q16)
async def s5q16_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s5q16=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[16].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FifthSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FifthSection.s5q17)
async def s5q17_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s5q17=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[7].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FifthSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FifthSection.s5q18)
async def s5q18_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s5q18=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[18].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FifthSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FifthSection.s5q19)
async def s5q19_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s5q19=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[19].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await FifthSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FifthSection.s5q20)
async def s5q20_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s5q20=int(message.text))
            
            section = Section.objects.filter(type=SectionNumberChoices.FIVE).first()
            total_questions = section.total_questions
            
            five_section_result = (total_ball // total_questions) * 0.1
            await state.update_data(five_section_result=five_section_result)
            
            await message.answer(f'<strong>Вы оценили первый {section.title}</strong>\n'
                     f'Результат {five_section_result}% из 10% общего\n\n'
                     f'Чтобы продолжить оценивание и перейти к следующему разделу, нажмите на /next')
            
            await FifthSection.next()
            
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")    
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=FifthSection.final_s5)
async def final_s5_handler(message: types.Message, state: FSMContext):
    if message.text == "/next":
        total_ball = 0
        
        section = Section.objects.filter(type=SectionNumberChoices.SIX).first()
        questions = list(section.section_questions.order_by('order'))
        await state.update_data(questions=questions, total_ball=total_ball)
        
        await message.answer(f'<strong>Вы перешли на другую секцию</strong>')
        await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[0].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
        
        await SixSection.s6q1.set()
    else:
        await message.answer("Нажмите /next чтобы продолжить")


@dp.message_handler(state=SixSection.s6q1)
async def s6q1_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s6q1=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[1].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await SixSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SixSection.s6q2)
async def s6q2_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s6q2=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[2].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await SixSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SixSection.s6q3)
async def s6q3_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s6q3=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[3].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await SixSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SixSection.s6q4)
async def s6q4_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s6q4=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[4].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await SixSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SixSection.s6q5)
async def s6q5_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s6q5=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[5].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await SixSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SixSection.s6q6)
async def s6q6_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s6q6=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[6].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await SixSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        

@dp.message_handler(state=SixSection.s6q7)
async def s6q7_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s6q7=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[7].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await SixSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SixSection.s6q8)
async def s6q8_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s6q8=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[8].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await SixSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SixSection.s6q9)
async def s6q9_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s6q9=int(message.text))
            
            section = Section.objects.filter(type=SectionNumberChoices.SIX).first()
            total_questions = section.total_questions
            
            six_section_result = (total_ball // total_questions) * 0.1
            await state.update_data(six_section_result=six_section_result)
            
            await message.answer(f'<strong>Вы оценили первый {section.title}</strong>\n'
                     f'Результат {six_section_result}% из 10% общего\n\n'
                     f'Чтобы продолжить оценивание и перейти к следующему разделу, нажмите на /next')
            
            await SixSection.next()
            
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")    
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
@dp.message_handler(state=SixSection.final_s6)
async def final_s6_handler(message: types.Message, state: FSMContext):
    if message.text == "/next":
        total_ball = 0
        
        section = Section.objects.filter(type=SectionNumberChoices.SEVEN).first()
        questions = list(section.section_questions.order_by('order'))
        await state.update_data(questions=questions, total_ball=total_ball)
        
        await message.answer(f'<strong>Вы перешли на другую секцию</strong>')
        await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[0].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
        
        await SevenSection.s7q1.set()
    else:
        await message.answer("Нажмите /next чтобы продолжить")
        

@dp.message_handler(state=SevenSection.s7q1)
async def s7q1_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s7q1=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[1].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await SevenSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SevenSection.s7q2)
async def s7q2_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s7q2=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[2].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await SevenSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SevenSection.s7q3)
async def s7q3_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s7q3=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[3].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await SevenSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SevenSection.s7q4)
async def s7q4_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s7q4=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[4].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await SevenSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SevenSection.s7q5)
async def s7q5_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s7q5=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[5].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await SevenSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SevenSection.s7q6)
async def s7q6_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s7q6=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[6].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await SevenSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SevenSection.s7q7)
async def s7q7_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s7q7=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[7].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await SevenSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SevenSection.s7q8)
async def s7q8_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s7q8=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[8].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await SevenSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SevenSection.s7q9)
async def s7q9_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s7q9=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[9].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await SevenSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
        
@dp.message_handler(state=SevenSection.s7q10)
async def s7q10_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            questions = data.get('questions', [])
            total_ball = data.get('total_ball', 0)
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s7q10=int(message.text))
            
            await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[10].title}\n\n'
                            f'Оцените, пожалуйста, от 0 до 100 баллов.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
            await SevenSection.next()
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SevenSection.s7q11)
async def s7q11_handler(message: types.Message, state: FSMContext):
    try:
        if 0 <= int(message.text) <= 100:
            data = await state.get_data()
            total_ball = data.get('total_ball')
            total_ball += int(message.text)
            await state.update_data(total_ball=total_ball, s7q11=int(message.text))
            
            section = Section.objects.filter(type=SectionNumberChoices.SEVEN).first()
            total_questions = section.total_questions
            
            seven_section_result = (total_ball // total_questions) * 0.1
            await state.update_data(seven_section_result=seven_section_result)
            
            branch = data.get("branch")
            name = data.get("name")
            await message.answer(f'Branch name: {branch}\n'
                                 f'Your Name {name}\n')
            
            first_section = data.get("first_section_result")
            second_section = data.get("second_section_result")
            third_section = data.get("third_section_result")
            four_section = data.get("four_section_result")
            five_section = data.get("five_section_result")
            six_section = data.get("six_section_result")
            seven_section = seven_section_result
            # 17.1, 17.1, 11.1 , 6.2, 5.2 , 4.2 , 
            await message.answer(f'First Result {first_section}%\n'
                  f'Second Result {second_section}%\n'
                  f'Third Result {third_section}%\n'  
                  f'Four Result {four_section}%\n'
                  f'Five Result {five_section}%\n'
                  f'Six Result {six_section}%\n'
                  f'Seven Result {seven_section}%\n')
            total_score = sum([first_section, second_section, third_section, four_section, five_section, six_section, seven_section])
            await message.answer(f'Total score:{total_score}')
            await state.update_data(total_score=total_score)
            
            await message.answer(f'<strong>Вы оценили первый {section.title}</strong>\n'
                     f'Результат {seven_section_result}% из 10% общего\n\n'
                     f'Чтобы продолжить оценивание и перейти к следующему разделу, нажмите на /next')
                                    
            await SevenSection.next()
            
        else:
            await message.answer("Ошибка! Баллы должны быть между 0 и 100.")    
    except ValueError:
        await message.answer("Ошибка! Пожалуйста, введите число.")
        
        
@dp.message_handler(state=SevenSection.finale_s7)
async def final_s7_handler(message: types.Message, state: FSMContext):
    if message.text == "/next":
        total_ball = 0
        
        section = Section.objects.filter(type=SectionNumberChoices.TEXT_QUESTION).first()
        questions = list(section.section_questions.order_by('order'))
        await state.update_data(questions=questions, total_ball=total_ball)
        
        await message.answer(f'<strong>Вы перешли на другую секцию</strong>')
        await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[0].title}\n\n'
                            f'Оценитe, написав комментарий.',
                            reply_markup=get_back_keyboard(),
                            parse_mode='HTML')
        
        await TextQuestionSection.t1q1.set()
    else:
        await message.answer("Нажмите /next чтобы продолжить")
        
        
@dp.message_handler(state=TextQuestionSection.t1q1)
async def t1q1_handler(message: types.Message, state: FSMContext):

    data = await state.get_data()
    questions = data.get('questions', [])
    await state.update_data(t1q1=message.text)
    
    await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[1].title}\n\n'
                    f'Оцените,написав комментарий..',
                    reply_markup=get_back_keyboard(),
                    parse_mode='HTML')
    await TextQuestionSection.next()

        
@dp.message_handler(state=TextQuestionSection.t1q2)
async def t1q2_handler(message: types.Message, state: FSMContext):
    section = Section.objects.filter(id=9).first()
    questions = list(section.section_questions.order_by('order'))
    await state.update_data(questions=questions, t1q2=message.text)

    await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[0].title}\n\n'
                    f'Оцените,написав комментарий..',
                    reply_markup=get_back_keyboard(),
                    parse_mode='HTML')
    await TextQuestionSection.next()
        
        
@dp.message_handler(state=TextQuestionSection.t1q3)
async def t1q3_handler(message: types.Message, state: FSMContext):
    data = await state.get_data()
    await state.update_data(t1q3=message.text)
    questions = data.get('questions', [])

    await message.answer(f'<strong>{questions[0].section.title}</strong>\n\n{questions[1].title}\n\n'
                    f'Оцените,написав комментарий..',
                    reply_markup=get_back_keyboard(),
                    parse_mode='HTML')
    
    await TextQuestionSection.next()
    
    
@dp.message_handler(state=TextQuestionSection.t1q4)
async def t1q4_handler(message: types.Message, state: FSMContext):
    await state.update_data(t1q4=message.text)
    await message.answer(f'Напиши что угодно чтобы сгенерировать excel file')        
    await TextQuestionSection.next()
    
      
@dp.message_handler(state=TextQuestionSection.generate_excel)
async def generate_excel_handler(message: types.Message, state: FSMContext):
    data = await state.get_data()
    print("works")
    wb = Workbook()
    ws = wb.active
    ws.title = "Main CHEK-List"
    
    fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")

    
    headers = ["Название Магазина", "Время", "Дата проверки", "Проверяющий"]
    cell = ws.cell(row=1, column=1, value=headers[0])
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    cell = ws.cell(row=1, column=2, value=headers[1])
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    cell = ws.cell(row=1, column=3, value=headers[2])
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    cell = ws.cell(row=1, column=4, value=headers[3])
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

    ws.cell(row=2, column=1, value=data['branch']).alignment = Alignment(horizontal='center')
    ws.cell(row=2, column=2, value=data['time']).alignment = Alignment(horizontal='center')
    ws.cell(row=2, column=3, value=data['date']).alignment = Alignment(horizontal='center')
    ws.cell(row=2, column=4, value=data['name']).alignment = Alignment(horizontal='center')
    
    
    # First section
    section = Section.objects.filter(type=SectionNumberChoices.FIRST).first()
    questions = list(section.section_questions.order_by('order'))
    
    first_section_result = data.get('first_section_result')

    cell = ws.cell(row=4, column=1, value=section.title)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')
    ws.cell(row=4, column=2, value="вес").alignment = Alignment(horizontal='center')
    ws.cell(row=4, column=3, value="процент").alignment = Alignment(horizontal='center')
    ws.cell(row=4, column=4, value="балл").alignment = Alignment(horizontal='center')
    
    question_s1 = [q.title for q in questions]
        
    first_result = [
        data.get(f's1q{i+1}', '') for i in range(len(questions))
    ]

    for idx, question in enumerate(question_s1, start=5):
        ws.cell(row=idx, column=1, value=question).alignment = Alignment(horizontal='center')
        ws.cell(row=idx, column=2, value=100) .alignment = Alignment(horizontal='center')
        ws.cell(row=idx, column=4, value=first_result[idx-5]).alignment = Alignment(horizontal='center')
        
    ws.cell(row=19, column=3, value=first_section_result).alignment = Alignment(horizontal='center')
    
    # second section
    section = Section.objects.filter(type=SectionNumberChoices.SECOND).first()
    questions = list(section.section_questions.order_by('order'))
    
    second_section_result = data.get('second_section_result') 
    
    cell = ws.cell(row=20, column=1, value=section.title)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")  
    cell.alignment = Alignment(horizontal='center')  
    ws.cell(row=20, column=2, value="вес").alignment = Alignment(horizontal='center')
    ws.cell(row=20, column=3, value="процент").alignment = Alignment(horizontal='center')
    ws.cell(row=20, column=4, value="балл").alignment = Alignment(horizontal='center')
    
    question_s2 = [q.title for q in questions]
    
    second_result = [
        data.get(f's2q{i+1}', '') for i in range(len(questions))
    ]

    for idx, question in enumerate(question_s2, start=21):
        ws.cell(row=idx, column=1, value=question).alignment = Alignment(horizontal='center')
        ws.cell(row=idx, column=2, value=100).alignment = Alignment(horizontal='center')  
        ws.cell(row=idx, column=4, value=second_result[idx-21]).alignment = Alignment(horizontal='center')
        
    ws.cell(row=34, column=3, value=second_section_result).alignment = Alignment(horizontal='center')
        
    # third section
    section = Section.objects.filter(type=SectionNumberChoices.THIRD).first()
    questions = list(section.section_questions.order_by('order'))
    
    third_section_result = data.get('third_section_result')
    
    cell = ws.cell(row=35, column=1, value=section.title)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')
    ws.cell(row=35, column=2, value="вес").alignment = Alignment(horizontal='center')
    ws.cell(row=35, column=3, value="процент").alignment = Alignment(horizontal='center')
    ws.cell(row=35, column=4, value="балл").alignment = Alignment(horizontal='center')

    question_s3 = [q.title for q in questions]
    
    third_result = [
        data.get(f's3q{i+1}', '') for i in range(len(questions))
    ]
    
    for idx, question in enumerate(question_s3, start=36):
        ws.cell(row=idx, column=1, value=question).alignment = Alignment(horizontal='center')
        ws.cell(row=idx, column=2, value=100).alignment = Alignment(horizontal='center') 
        ws.cell(row=idx, column=4, value=third_result[idx-36]).alignment = Alignment(horizontal='center')
        
    ws.cell(row=57, column=3, value=third_section_result).alignment = Alignment(horizontal='center')
        
    
    # four section
    section = Section.objects.filter(type=SectionNumberChoices.FOUR).first()
    questions = list(section.section_questions.order_by('order'))
    four_section_result = data.get('four_section_result')
    
    cell = ws.cell(row=58, column=1, value=section.title)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')
    ws.cell(row=58, column=2, value="вес").alignment = Alignment(horizontal='center')
    ws.cell(row=58, column=3, value="процент").alignment = Alignment(horizontal='center')
    ws.cell(row=58, column=4, value="балл").alignment = Alignment(horizontal='center')

    question_s4 = [q.title for q in questions]
    
    four_result = [
        data.get(f's4q{i+1}', '') for i in range(len(questions))
    ]
    
    for idx, question in enumerate(question_s4, start=59):
        ws.cell(row=idx, column=1, value=question).alignment = Alignment(horizontal='center')
        ws.cell(row=idx, column=2, value=100).alignment = Alignment(horizontal='center')  
        ws.cell(row=idx, column=4, value=four_result[idx-59]).alignment = Alignment(horizontal='center')
        
    ws.cell(row=73, column=3, value=four_section_result).alignment = Alignment(horizontal='center')
         
    
    # five section
    section = Section.objects.filter(type=SectionNumberChoices.FIVE).first()
    questions = list(section.section_questions.order_by('order'))
    
    five_section_result = data.get('five_section_result')
    
    cell = ws.cell(row=74, column=1, value=section.title)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')
    ws.cell(row=74, column=2, value="вес").alignment = Alignment(horizontal='center')
    ws.cell(row=74, column=3, value="процент").alignment = Alignment(horizontal='center')
    ws.cell(row=74, column=4, value="балл").alignment = Alignment(horizontal='center')

    question_s5 = [q.title for q in questions]
    
    five_result = [
        data.get(f's5q{i+1}', '') for i in range(len(questions))
    ]
    
    for idx, question in enumerate(question_s5, start=75):
        ws.cell(row=idx, column=1, value=question).alignment = Alignment(horizontal='center')
        ws.cell(row=idx, column=2, value=100).alignment = Alignment(horizontal='center')  
        ws.cell(row=idx, column=4, value=five_result[idx-75]).alignment = Alignment(horizontal='center')
        
    ws.cell(row=95, column=3, value=five_section_result)
        
    
    # six section
    section = Section.objects.filter(type=SectionNumberChoices.SIX).first()
    questions = list(section.section_questions.order_by('order'))
    
    six_section_result = data.get('six_section_result')
    
    cell = ws.cell(row=96, column=1, value=section.title)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')
    ws.cell(row=96, column=2, value="вес").alignment = Alignment(horizontal='center')
    ws.cell(row=96, column=3, value="процент").alignment = Alignment(horizontal='center')
    ws.cell(row=96, column=4, value="балл").alignment = Alignment(horizontal='center')

    question_s6 = [q.title for q in questions]
    
    six_result = [
        data.get(f's6q{i+1}','') for i in range(len(questions))
    ]
    
    for idx, question in enumerate(question_s6, start=97):
        ws.cell(row=idx, column=1, value=question).alignment = Alignment(horizontal='center')
        ws.cell(row=idx, column=2, value=100).alignment = Alignment(horizontal='center')  
        ws.cell(row=idx, column=4, value=six_result[idx-97]).alignment = Alignment(horizontal='center')
        
    ws.cell(row=106, column=3, value=six_section_result).alignment = Alignment(horizontal='center')
    
    
    # seven section
    section = Section.objects.filter(type=SectionNumberChoices.SEVEN).first()
    questions = list(section.section_questions.order_by('order'))
    seven_section_result = data.get('seven_section_result')
    
    cell = ws.cell(row=107, column=1, value=section.title)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')
    ws.cell(row=107, column=2, value="вес").alignment = Alignment(horizontal='center')
    ws.cell(row=107, column=3, value="процент").alignment = Alignment(horizontal='center')
    ws.cell(row=107, column=4, value="балл").alignment = Alignment(horizontal='center')

    question_s7 = [q.title for q in questions]
    
    seven_result = [
       data.get(f's7q{i+1}','') for i in range(len(questions))
    ]
    
    for idx, question in enumerate(question_s7, start=107):
        ws.cell(row=idx, column=1, value=question).alignment = Alignment(horizontal='center')
        ws.cell(row=idx, column=2, value=100).alignment = Alignment(horizontal='center')  
        ws.cell(row=idx, column=4, value=seven_result[idx-107]).alignment = Alignment(horizontal='center')
        
    ws.cell(row=118, column=3, value=seven_section_result).alignment = Alignment(horizontal='center')
    
    total_score = data.get('total_score', 0)
    # total result  
    cell = ws.cell(row=119, column=1, value="Результат")
    cell = ws.cell(row=119, column=2, value="10100")
    cell = ws.cell(row=119, column=3, value="weight/total_Ball")
    cell = ws.cell(row=119, column=4, value="total ball")
    cell = ws.cell(row=120, column=1, value=total_score)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')
    
        
    
    # first text section
    section = Section.objects.filter(type=SectionNumberChoices.TEXT_QUESTION).first()
    questions = list(section.section_questions.order_by('order'))
    
    cell = ws.cell(row=122, column=1, value=section.title)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')
    ws.cell(row=122, column=2, value="коментарий").alignment = Alignment(horizontal='center')
    
    ws.cell(row=123, column=1, value=questions[0].title).alignment = Alignment(horizontal='center')
    ws.cell(row=123, column=2, value=data.get('t1q1')).alignment = Alignment(horizontal='center')
    ws.cell(row=124, column=1, value=questions[1].title).alignment = Alignment(horizontal='center')
    ws.cell(row=124, column=2, value=data.get('t1q2')).alignment = Alignment(horizontal='center')
        
        
    # second text section
    section = Section.objects.filter(id=9).first()
    questions = list(section.section_questions.order_by('order'))
    
    cell = ws.cell(row=125, column=1, value=section.title)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')
    ws.cell(row=125, column=2, value="коментарий").alignment = Alignment(horizontal='center')

    ws.cell(row=126, column=1, value=questions[0].title).alignment = Alignment(horizontal='center')
    ws.cell(row=126, column=2, value=data.get('t1q3')).alignment = Alignment(horizontal='center')
    ws.cell(row=127, column=1, value=questions[1].title).alignment = Alignment(horizontal='center')
    ws.cell(row=127, column=2, value=data.get('t1q4')).alignment = Alignment(horizontal='center')

    wb.save("checklist.xlsx")

    await message.answer("Файл создан! Вот ваш файл:")
    await message.answer_document(types.InputFile("checklist.xlsx"))

    await state.finish()
